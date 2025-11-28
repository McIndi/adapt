from __future__ import annotations
import logging
from adapt.cache import get_cache, set_cache, invalidate_cache

from pathlib import Path
from typing import Any, Sequence

from openpyxl import load_workbook

from .base import ResourceDescriptor
from .dataset_plugin import DatasetPlugin


logger = logging.getLogger(__name__)


class ExcelPlugin(DatasetPlugin):
    @property
    def resource_type(self) -> str:
        """Return the resource type string."""
        return "excel"

    def detect(self, path: Path) -> bool:
        """Detect if the path is an Excel file.

        Args:
            path: The file path to check.

        Returns:
            True if the file has .xlsx extension, False otherwise.
        """
        return path.suffix.lower() == ".xlsx"

    def load(self, path: Path) -> Sequence[ResourceDescriptor]:
        """Load Excel file and create descriptors for each worksheet.

        Args:
            path: The path to the Excel file.

        Returns:
            A sequence of ResourceDescriptors, one for each worksheet.
        """
        logger.debug(f"Loading Excel file: {path}")
        workbook = load_workbook(path, read_only=True, data_only=True)
        descriptors = []
        try:
            for sheet_name in workbook.sheetnames:
                logger.debug(f"Processing sheet: {sheet_name}")
                sheet = workbook[sheet_name]
                rows = sheet.iter_rows(values_only=True)
                header_row = list(next(rows, []))
                sample_row = list(next(rows, []))
                from .dataset_plugin import _ensure_header
                normalized = _ensure_header(header_row)
                normalized_sample = [str(cell) if cell is not None else None for cell in sample_row]
                descriptor = ResourceDescriptor(path=path, resource_type=self.resource_type)
                descriptor.metadata["header"] = normalized
                descriptor.metadata["sample_row"] = normalized_sample
                descriptor.metadata["primary_key"] = "_row_id"
                descriptor.metadata["sub_namespace"] = sheet_name
                descriptors.append(descriptor)
        finally:
            workbook.close()
        logger.info(f"Loaded {len(descriptors)} worksheets from Excel file: {path}")
        return descriptors

    def _read_raw_rows(self, resource: ResourceDescriptor) -> list[list[str]]:
        """Read raw rows from the Excel worksheet.

        Args:
            resource: The resource descriptor.

        Returns:
            A list of rows as lists of strings.
        """
        logger.debug(f"Reading raw rows from Excel sheet: {resource.metadata.get('sub_namespace', '')}")
        cache_key = f"data:{resource.path}:{resource.metadata.get('sub_namespace', '')}"
        cached = get_cache(cache_key, str(resource.path))
        if cached:
            logger.debug(f"Cache hit for Excel data: {resource.path}")
            return cached
        logger.debug(f"Cache miss, reading from Excel file: {resource.path}")
        sub_namespace = resource.metadata["sub_namespace"]
        rows = []
        workbook = load_workbook(resource.path, read_only=True, data_only=True)
        try:
            sheet = workbook[sub_namespace]
            data_rows = list(sheet.iter_rows(values_only=True))
            for row in data_rows[1:]:  # Skip header
                rows.append([str(cell) if cell is not None else "" for cell in row])
        finally:
            workbook.close()
        set_cache(cache_key, rows, ttl_seconds=300, resource=str(resource.path))  # 5 min TTL
        return rows

    def _write_rows(self, resource: ResourceDescriptor, rows: list[dict[str, Any]], header: list[str]) -> None:
        """Write rows to the Excel worksheet.

        Args:
            resource: The resource descriptor.
            rows: The rows to write.
            header: The column headers.
        """
        logger.info(f"Writing rows to Excel worksheet: {resource.metadata.get('sub_namespace', '')}")
        sub_namespace = resource.metadata["sub_namespace"]
        # Write back atomically
        import tempfile
        import os
        workbook = load_workbook(resource.path)
        try:
            sheet = workbook[sub_namespace]
            # Clear existing data
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.value = None
            # Write new data
            for row_idx, row_data in enumerate(rows, start=2):
                for col_idx, col_name in enumerate(header):
                    sheet.cell(row=row_idx, column=col_idx+1).value = row_data.get(col_name, "")
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_fh:
                workbook.save(tmp_fh.name)
                tmp_path = tmp_fh.name
        finally:
            workbook.close()

        os.replace(tmp_path, resource.path)
        # Invalidate cache after mutation
        invalidate_cache(str(resource.path))
        logger.debug(f"Successfully wrote {len(rows)} rows to Excel file: {resource.path}")
    
