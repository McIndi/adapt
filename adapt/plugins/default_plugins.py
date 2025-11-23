from __future__ import annotations

import csv
from pathlib import Path
from typing import Any, Sequence

from fastapi import Request, Depends
from fastapi.routing import APIRouter
from openpyxl import load_workbook

from .base import Plugin, ResourceDescriptor, PluginContext


def _guess_type(value: str | None) -> str:
    if value is None:
        return "string"

    candidate = str(value).strip()
    if not candidate:
        return "string"

    try:
        int(candidate)
        return "integer"
    except ValueError:
        pass

    try:
        float(candidate)
        return "number"
    except ValueError:
        pass

    lower = candidate.lower()
    if lower in {"true", "false"}:
        return "boolean"

    return "string"


def _ensure_header(header: Sequence[str | None]) -> list[str]:
    return [str(col).strip() if col else f"column_{idx + 1}" for idx, col in enumerate(header)]


def _build_columns(header: Sequence[str], sample: Sequence[str | None]) -> dict[str, dict[str, str]]:
    columns: dict[str, dict[str, str]] = {}
    for idx, column in enumerate(header):
        sample_value = sample[idx] if idx < len(sample) else None
        columns[column] = {"type": _guess_type(sample_value)}
    return columns


class DatasetPlugin(Plugin):
    def load(self, path: Path) -> ResourceDescriptor:
        header, sample = self._get_header_and_sample(path)
        descriptor = ResourceDescriptor(path=path, resource_type=self.resource_type)
        descriptor.metadata["header"] = header
        descriptor.metadata["sample_row"] = sample
        descriptor.metadata["primary_key"] = "_row_id"
        return descriptor

    def schema(self, resource: ResourceDescriptor) -> dict[str, Any]:
        if resource.schema_path and resource.schema_path.exists():
            import json
            with resource.schema_path.open() as f:
                return json.load(f)
        else:
            header = resource.metadata.get("header", [])
            sample = resource.metadata.get("sample_row", [])
            columns = _build_columns(header, sample)
            return {
                "type": "object",
                "name": resource.path.stem,
                "primary_key": resource.metadata.get("primary_key"),
                "columns": columns,
            }

    def read(self, resource: ResourceDescriptor, request: Request) -> Sequence[dict[str, Any]]:
        header = resource.metadata.get("header", [])
        schema = self.schema(resource)
        columns = schema.get("columns", {})
        raw_rows = self._read_raw_rows(resource)
        rows = []
        for row_id, row in enumerate(raw_rows, start=1):
            row_dict = {"_row_id": row_id}
            for idx, value in enumerate(row):
                if idx < len(header):
                    col_name = header[idx]
                    col_schema = columns.get(col_name, {})
                    col_type = col_schema.get("type", "string")
                    row_dict[col_name] = self._cast_value(value, col_type)
            rows.append(row_dict)
        return rows

    def _cast_value(self, value: str, col_type: str) -> Any:
        if col_type == "integer":
            try:
                return int(value)
            except ValueError:
                return value
        elif col_type == "number":
            try:
                return float(value)
            except ValueError:
                return value
        elif col_type == "boolean":
            lower = value.lower()
            if lower in ("true", "1", "yes"):
                return True
            elif lower in ("false", "0", "no"):
                return False
            return value
        else:
            return value

    def write(self, resource: ResourceDescriptor, data: Any, request: Request, context: PluginContext) -> dict[str, Any]:
        action = data.get("action")
        payload = data.get("data", [])

        # Read existing data
        existing_rows = list(self.read(resource, request))
        header = resource.metadata.get("header", [])

        if action == "create":
            # Append new rows
            for new_row in payload:
                row_id = len(existing_rows) + 1
                row_dict = {"_row_id": row_id}
                for col in header:
                    row_dict[col] = new_row.get(col, "")
                existing_rows.append(row_dict)
        elif action == "update":
            # Update existing row
            row_id = int(payload.get("_row_id"))
            for row in existing_rows:
                if row["_row_id"] == row_id:
                    for col in header:
                        if col in payload:
                            row[col] = payload[col]
                    break
        elif action == "delete":
            # Remove row
            row_id = int(payload.get("_row_id"))
            existing_rows = [row for row in existing_rows if row["_row_id"] != row_id]
            # Reassign row_ids
            for idx, row in enumerate(existing_rows, start=1):
                row["_row_id"] = idx

        # Write back
        self._write_rows(resource, existing_rows, header)

        return {"success": True}

    def routes(self, resource: ResourceDescriptor) -> Sequence[APIRouter]:
        router = APIRouter()

        @router.get("/")
        def read_all(request: Request):
            return self.read(resource, request)

        @router.post("/")
        def create(data: dict, request: Request):
            context = PluginContext(
                engine=request.app.state.db_engine,
                root=request.app.state.config.root,
                readonly=request.app.state.config.readonly
            )
            return self.write(resource, data, request, context)

        @router.patch("/")
        def update(data: dict, request: Request):
            context = PluginContext(
                engine=request.app.state.db_engine,
                root=request.app.state.config.root,
                readonly=request.app.state.config.readonly
            )
            return self.write(resource, data, request, context)

        @router.delete("/")
        def delete(data: dict, request: Request):
            context = PluginContext(
                engine=request.app.state.db_engine,
                root=request.app.state.config.root,
                readonly=request.app.state.config.readonly
            )
            return self.write(resource, data, request, context)

        return [router]

    @property
    def resource_type(self) -> str:
        raise NotImplementedError

    def _get_header_and_sample(self, path: Path) -> tuple[list[str], list[str | None]]:
        raise NotImplementedError

    def _read_raw_rows(self, resource: ResourceDescriptor) -> list[list[str]]:
        raise NotImplementedError

    def _write_rows(self, resource: ResourceDescriptor, rows: list[dict[str, Any]], header: list[str]) -> None:
        raise NotImplementedError


class CsvPlugin(DatasetPlugin):
    @property
    def resource_type(self) -> str:
        return "csv"

    def detect(self, path: Path) -> bool:
        return path.suffix.lower() == ".csv"

    def _get_header_and_sample(self, path: Path) -> tuple[list[str], list[str | None]]:
        with path.open(newline="", encoding="utf-8") as fh:
            reader = csv.reader(fh)
            header = next(reader, [])
            sample = next(reader, [])
        normalized = _ensure_header(header)
        return normalized, sample

    def _read_raw_rows(self, resource: ResourceDescriptor) -> list[list[str]]:
        with resource.path.open(newline="", encoding="utf-8") as fh:
            reader = csv.reader(fh)
            next(reader, None)  # Skip header
            return list(reader)

    def _write_rows(self, resource: ResourceDescriptor, rows: list[dict[str, Any]], header: list[str]) -> None:
        # Write back atomically
        import tempfile
        import os
        with tempfile.NamedTemporaryFile(mode='w', suffix='.csv', delete=False, newline='', encoding='utf-8') as tmp_fh:
            writer = csv.writer(tmp_fh)
            writer.writerow(header)
            for row in rows:
                writer.writerow([row.get(col, "") for col in header])
            tmp_path = tmp_fh.name

        os.replace(tmp_path, resource.path)


class ExcelPlugin(DatasetPlugin):
    @property
    def resource_type(self) -> str:
        return "excel"

    def detect(self, path: Path) -> bool:
        return path.suffix.lower() == ".xlsx"

    def _get_header_and_sample(self, path: Path) -> tuple[list[str], list[str | None]]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            header_row = list(next(rows, []))
            sample_row = list(next(rows, []))
        finally:
            workbook.close()
        normalized = _ensure_header(header_row)
        normalized_sample = [str(cell) if cell is not None else None for cell in sample_row]
        return normalized, normalized_sample

    def _read_raw_rows(self, resource: ResourceDescriptor) -> list[list[str]]:
        rows = []
        workbook = load_workbook(resource.path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            data_rows = list(sheet.iter_rows(values_only=True))
            for row in data_rows[1:]:  # Skip header
                rows.append([str(cell) if cell is not None else "" for cell in row])
        finally:
            workbook.close()
        return rows

    def _write_rows(self, resource: ResourceDescriptor, rows: list[dict[str, Any]], header: list[str]) -> None:
        # Write back atomically
        import tempfile
        import os
        workbook = load_workbook(resource.path)
        try:
            sheet = workbook.active
            # Clear existing data
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    cell.value = None
            # Write new data
            for row_idx, row_data in enumerate(rows, start=2):
                for col_idx, col_name in enumerate(header):
                    sheet.cell(row=row_idx, column=col_idx+1).value = row_data.get(col_name, "")
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_fh:
                workbook.save(tmp_fh.name)
                tmp_path = tmp_fh.name
        finally:
            workbook.close()

        os.replace(tmp_path, resource.path)


class PythonHandlerPlugin(Plugin):
    def detect(self, path: Path) -> bool:
        return path.suffix.lower() == ".py"

    def load(self, path: Path) -> ResourceDescriptor:
        descriptor = ResourceDescriptor(path=path, resource_type="python")
        return descriptor

    def schema(self, resource: ResourceDescriptor) -> dict[str, Any]:
        return {}  # No schema for Python handlers

    def read(self, resource: ResourceDescriptor, request: Request) -> Any:
        raise NotImplementedError("Python handlers do not support read operations")

    def write(self, resource: ResourceDescriptor, data: Any, request: Request, context: PluginContext) -> Any:
        raise NotImplementedError("Python handlers do not support write operations")

    def routes(self, resource: ResourceDescriptor) -> Sequence[APIRouter]:
        import importlib.util
        spec = importlib.util.spec_from_file_location(resource.path.stem, resource.path)
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        if hasattr(module, 'router') and isinstance(module.router, APIRouter):
            return [module.router]
        return []